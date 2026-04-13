"""
Excel-to-SQLite import service for the Budget Platform.

Reads all source Excel workbooks (Budget Sheet, Invoices, Sales Orders,
Credit Notes, Proposals, Master Data) and bulk-loads them into the database,
replacing any previously imported data on each run.
"""

from __future__ import annotations

import logging
import os
from datetime import datetime, date
from pathlib import Path
from typing import Any, List, Optional, Sequence

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.models import (
    BillingEntity,
    BudgetLine,
    BudgetMonthly,
    Client,
    CreditNote,
    Department,
    Invoice,
    Manager,
    Partner,
    PipelineEntry,
    Proposal,
    SalesOrder,
    ServiceCategory,
    TrueUpRemark,
    VarianceReason,
)

logger = logging.getLogger(__name__)

MONTH_LABELS = [
    "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25",
    "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26",
]

BATCH_SIZE = 500


# ── helpers ──────────────────────────────────────────────────────────────────


def _open_workbook(path: str | Path, read_only: bool = True):
    """Open an Excel workbook with openpyxl, raising a clear error on failure."""
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Workbook not found: {path}")
    return load_workbook(str(path), read_only=read_only, data_only=True)


def _cell(row: tuple, col: int) -> Any:
    """Return the value of a 1-based column index from an openpyxl row tuple."""
    idx = col - 1
    if idx < 0 or idx >= len(row):
        return None
    return row[idx].value if hasattr(row[idx], "value") else row[idx]


def _str(val: Any, max_len: int | None = None) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if s == "" or s.lower() == "none" or s == "#N/A" or s == "#REF!":
        return None
    if max_len:
        s = s[:max_len]
    return s


def _float(val: Any) -> Optional[float]:
    if val is None:
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def _int(val: Any) -> Optional[int]:
    if val is None:
        return None
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return None


def _parse_date(val: Any) -> Optional[date]:
    """Parse a date that may arrive as a datetime object or a string."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _invoice_month(dt: Optional[date]) -> Optional[str]:
    """Convert a date to 'Mon-YY' format, e.g. date(2025,4,15) -> 'Apr-25'."""
    if dt is None:
        return None
    return dt.strftime("%b-%y")


def _billing_entity_from_path(filepath: str) -> str:
    """Derive 'IBA' or 'NMSA' from the filename."""
    name = Path(filepath).stem.upper()
    if "NMSA" in name:
        return "NMSA"
    return "IBA"


def _truncate(db: Session, model) -> None:
    """Delete all rows from the given table."""
    db.query(model).delete()
    db.flush()


def _flush_batch(db: Session, batch: list) -> None:
    if batch:
        db.bulk_save_objects(batch)
        db.flush()
        batch.clear()


def _generate_unique_code(so_number: Any, service_category: Any) -> Optional[str]:
    so = _str(so_number)
    svc = _str(service_category)
    if so and svc:
        return f"{so}|{svc}"
    if so:
        return so
    return None


# ── Master Data ──────────────────────────────────────────────────────────────


def import_master_data(db: Session, base_path: str) -> None:
    """Import master dropdown values from the 'Master' sheet of the Budget Sheet."""
    logger.info("Importing master data …")
    wb_path = os.path.join(base_path, "Budget Sheet 2025-26.xlsx")
    wb = _open_workbook(wb_path)
    try:
        ws = wb["Master"]
        rows = list(ws.iter_rows(min_row=2))  # skip header row 1
    finally:
        wb.close()

    col_map = {
        1: ("service_categories", set()),
        2: ("departments", set()),
        3: ("managers", set()),
        4: ("partners", set()),
        5: ("billing_entities", set()),
        7: ("clients", set()),
        19: ("variance_reasons", set()),
        23: ("true_up_remarks", set()),
    }

    for row in rows:
        for col_idx, (_, vals) in col_map.items():
            v = _str(_cell(row, col_idx))
            if v:
                vals.add(v)

    _truncate(db, ServiceCategory)
    _truncate(db, Department)
    _truncate(db, Manager)
    _truncate(db, Partner)
    _truncate(db, BillingEntity)
    _truncate(db, Client)
    _truncate(db, VarianceReason)
    _truncate(db, TrueUpRemark)

    dept_id_map: dict[str, int] = {}

    for name in sorted(col_map[2][1]):
        obj = Department(name=name)
        db.add(obj)
    db.flush()
    for d in db.query(Department).all():
        dept_id_map[d.name] = d.id

    for name in sorted(col_map[1][1]):
        sc = ServiceCategory(name=name, department_id=None)
        db.add(sc)
    db.flush()

    for name in sorted(col_map[3][1]):
        db.add(Manager(name=name))
    for name in sorted(col_map[4][1]):
        db.add(Partner(name=name))
    for name in sorted(col_map[5][1]):
        db.add(BillingEntity(name=name))
    for name in sorted(col_map[7][1]):
        db.add(Client(name=name))
    for reason in sorted(col_map[19][1]):
        db.add(VarianceReason(reason=reason))
    for remark in sorted(col_map[23][1]):
        db.add(TrueUpRemark(remark=remark))

    db.flush()
    logger.info(
        "Master data imported: %d service categories, %d departments, "
        "%d managers, %d partners, %d billing entities, %d clients, "
        "%d variance reasons, %d true-up remarks",
        len(col_map[1][1]), len(col_map[2][1]), len(col_map[3][1]),
        len(col_map[4][1]), len(col_map[5][1]), len(col_map[7][1]),
        len(col_map[19][1]), len(col_map[23][1]),
    )


# ── Budget Lines + Monthly ───────────────────────────────────────────────────


def import_budget_sheet(db: Session, base_path: str) -> None:
    """Import from 'Combined Budget' sheet of Budget Sheet 2025-26.xlsx."""
    logger.info("Importing budget sheet …")
    wb_path = os.path.join(base_path, "Budget Sheet 2025-26.xlsx")
    wb = _open_workbook(wb_path)
    try:
        ws = wb["Combined Budget"]
        rows = list(ws.iter_rows(min_row=5))  # header is row 4, data from row 5
    finally:
        wb.close()

    _truncate(db, BudgetMonthly)
    _truncate(db, BudgetLine)

    batch_lines: list[BudgetLine] = []
    count = 0

    for row in rows:
        serial = _int(_cell(row, 2))
        client = _str(_cell(row, 5))
        if serial is None and client is None:
            continue  # skip empty / summary rows

        bl = BudgetLine(
            serial_no=serial,
            quotation_no=_str(_cell(row, 3), 100),
            sales_order_no=_str(_cell(row, 4), 100),
            client_name=_str(_cell(row, 5), 500),
            billing_type=_str(_cell(row, 6), 50),
            billing_entity=_str(_cell(row, 7), 50),
            partner=_str(_cell(row, 8), 200),
            manager=_str(_cell(row, 9), 200),
            department=_str(_cell(row, 10), 100),
            service_category=_str(_cell(row, 11), 300),
            service_description=_str(_cell(row, 12)),
            billing_frequency=_str(_cell(row, 13), 50),
            no_of_billing=_int(_cell(row, 14)),
            currency=_str(_cell(row, 15), 10),
            exchange_rate=_float(_cell(row, 16)),
            budget_base=_str(_cell(row, 17), 50),
            loe_period_upto=_parse_date(_cell(row, 18)),
            existing_fees=_float(_cell(row, 19)),
            pct_increase=_float(_cell(row, 20)),
            increased_fees=_float(_cell(row, 21)),
            fee_for_sales_order=_float(_cell(row, 22)),
            amount_carried_forward=_float(_cell(row, 36)),
            check_value=_float(_cell(row, 37)),
            remarks=_str(_cell(row, 38)),
            unique_code_so=_generate_unique_code(
                _cell(row, 4), _cell(row, 11)
            ),
            unique_code_invoice=_generate_unique_code(
                _cell(row, 4), _cell(row, 11)
            ),
            sales_order_value=_float(_cell(row, 41)),
            variance=_float(_cell(row, 42)),
        )
        db.add(bl)
        db.flush()  # get bl.id

        # 12 monthly expected values: cols 23-34
        for mi in range(12):
            expected_col = 23 + mi
            # Actual/MTD/YTD/Reason/Remark blocks start at col 43 and repeat every 5 cols
            actual_base = 43 + mi * 5
            bm = BudgetMonthly(
                budget_line_id=bl.id,
                month=MONTH_LABELS[mi],
                month_index=mi,
                expected=_float(_cell(row, expected_col)),
                actual=_float(_cell(row, actual_base)),
                mtd_variance=_float(_cell(row, actual_base + 1)),
                ytd_variance=_float(_cell(row, actual_base + 2)),
                reason=_str(_cell(row, actual_base + 3), 300),
                remark=_str(_cell(row, actual_base + 4)),
            )
            db.add(bm)

        count += 1
        if count % BATCH_SIZE == 0:
            db.flush()

    db.flush()
    logger.info("Budget sheet imported: %d lines", count)


# ── Invoices ─────────────────────────────────────────────────────────────────


def _import_invoice_file(db: Session, filepath: str, batch: list) -> int:
    """Parse a single invoice export file and append Invoice objects to *batch*."""
    entity = _billing_entity_from_path(filepath)
    wb = _open_workbook(filepath)
    try:
        ws = wb["Invoice"]
        rows = list(ws.iter_rows(min_row=2))  # header is row 1
    finally:
        wb.close()

    count = 0
    for row in rows:
        inv_date = _parse_date(_cell(row, 1))
        inv_number = _str(_cell(row, 2))
        if not inv_number:
            continue

        status = _str(_cell(row, 3), 50)
        is_void = status is not None and status.lower() == "void"
        raw_item_total = _float(_cell(row, 60))

        so_number = _str(_cell(row, 65))
        dept = _str(_cell(row, 167))
        item_name = _str(_cell(row, 55))

        inv = Invoice(
            invoice_date=inv_date,
            invoice_number=inv_number,
            invoice_status=status,
            customer_name=_str(_cell(row, 4), 500),
            billing_entity=entity,
            due_date=_parse_date(_cell(row, 9)),
            purchase_order=_str(_cell(row, 10), 100),
            currency_code=_str(_cell(row, 11), 10),
            exchange_rate=_float(_cell(row, 12)),
            template_name=_str(_cell(row, 15), 100),
            subtotal=_float(_cell(row, 24)),
            total=_float(_cell(row, 25)),
            balance=_float(_cell(row, 26)),
            item_name=_str(item_name, 300),
            item_desc=_str(_cell(row, 56)),
            quantity=_float(_cell(row, 57)),
            item_price=_float(_cell(row, 62)),
            item_total=raw_item_total,
            account=_str(_cell(row, 158), 100),
            department=_str(dept, 100),
            sales_order_number=_str(so_number, 100),
            unique_code=_generate_unique_code(so_number, item_name),
            invoice_month=_invoice_month(inv_date),
            is_voided=is_void,
            cleaned_item_total=0.0 if is_void else (raw_item_total or 0.0),
        )
        batch.append(inv)
        count += 1

    return count


def import_invoices(db: Session, base_path: str) -> None:
    """Import invoice exports from IBA and NMSA files."""
    logger.info("Importing invoices …")
    _truncate(db, Invoice)

    inv_dir = os.path.join(base_path, "Zoho Exports and Proposal", "Invoices")
    files = [
        os.path.join(inv_dir, "Inv_IBA.xlsx"),
        os.path.join(inv_dir, "Inv_NMSA.xlsx"),
    ]

    total = 0
    batch: list[Invoice] = []
    for fp in files:
        if not os.path.exists(fp):
            logger.warning("Invoice file not found, skipping: %s", fp)
            continue
        n = _import_invoice_file(db, fp, batch)
        total += n
        if len(batch) >= BATCH_SIZE:
            _flush_batch(db, batch)

    _flush_batch(db, batch)
    logger.info("Invoices imported: %d rows", total)


# ── Sales Orders ─────────────────────────────────────────────────────────────


def _import_so_file(db: Session, filepath: str, batch: list) -> int:
    entity = _billing_entity_from_path(filepath)
    wb = _open_workbook(filepath)
    try:
        ws = wb["Sales Order"]
        rows = list(ws.iter_rows(min_row=2))
    finally:
        wb.close()

    count = 0
    for row in rows:
        so_number = _str(_cell(row, 2))
        if not so_number:
            continue

        dept = _str(_cell(row, 96))
        item_name = _str(_cell(row, 18))

        so = SalesOrder(
            order_date=_parse_date(_cell(row, 1)),
            salesorder_number=so_number,
            status=_str(_cell(row, 3), 50),
            customer_name=_str(_cell(row, 4), 500),
            billing_entity=entity,
            gstin=_str(_cell(row, 8), 50),
            quotation_no=_str(_cell(row, 10), 100),
            currency_code=_str(_cell(row, 12), 10),
            exchange_rate=_float(_cell(row, 13)),
            item_name=_str(item_name, 300),
            item_desc=_str(_cell(row, 22)),
            quantity_ordered=_float(_cell(row, 23)),
            quantity_invoiced=_float(_cell(row, 24)),
            quantity_cancelled=_float(_cell(row, 25)),
            item_price=_float(_cell(row, 27)),
            item_total=_float(_cell(row, 56)),
            department=_str(dept, 100),
            sales_person=_str(_cell(row, 69), 200),
            unique_code=_generate_unique_code(so_number, item_name),
        )
        batch.append(so)
        count += 1

    return count


def import_sales_orders(db: Session, base_path: str) -> None:
    logger.info("Importing sales orders …")
    _truncate(db, SalesOrder)

    so_dir = os.path.join(base_path, "Zoho Exports and Proposal", "Sales Orders")
    files = [
        os.path.join(so_dir, "SO_IBA.xlsx"),
        os.path.join(so_dir, "SO_NMSA.xlsx"),
    ]

    total = 0
    batch: list[SalesOrder] = []
    for fp in files:
        if not os.path.exists(fp):
            logger.warning("SO file not found, skipping: %s", fp)
            continue
        n = _import_so_file(db, fp, batch)
        total += n
        if len(batch) >= BATCH_SIZE:
            _flush_batch(db, batch)

    _flush_batch(db, batch)
    logger.info("Sales orders imported: %d rows", total)


# ── Credit Notes ─────────────────────────────────────────────────────────────


def _build_invoice_po_lookup(db: Session) -> dict[str, str]:
    """Build a mapping of invoice_number -> purchase_order for PO lookups."""
    lookup: dict[str, str] = {}
    for inv_number, po in (
        db.query(Invoice.invoice_number, Invoice.purchase_order)
        .filter(Invoice.purchase_order.isnot(None))
        .distinct()
        .all()
    ):
        if inv_number and po:
            lookup[inv_number] = po
    return lookup


def _import_cn_file(filepath: str, po_lookup: dict[str, str], batch: list) -> int:
    entity = _billing_entity_from_path(filepath)
    wb = _open_workbook(filepath)
    try:
        ws = wb["CreditNotes"]
        rows = list(ws.iter_rows(min_row=2))
    finally:
        wb.close()

    count = 0
    for row in rows:
        cn_number = _str(_cell(row, 2))
        if not cn_number:
            continue

        raw_total = _float(_cell(row, 9))
        assoc_inv = _str(_cell(row, 5), 100)

        cn = CreditNote(
            cn_date=_parse_date(_cell(row, 1)),
            cn_number=cn_number,
            cn_status=_str(_cell(row, 3), 50),
            customer_name=_str(_cell(row, 4), 500),
            billing_entity=entity,
            associated_invoice_number=assoc_inv,
            item_name=_str(_cell(row, 6), 300),
            item_desc=_str(_cell(row, 7)),
            quantity=_float(_cell(row, 8)),
            item_total_original=raw_total,
            item_total_adjusted=(-raw_total) if raw_total is not None else None,
            account=_str(_cell(row, 10), 100),
            department=_str(_cell(row, 11), 100),
            purchase_order=po_lookup.get(assoc_inv) if assoc_inv else None,
        )
        batch.append(cn)
        count += 1

    return count


def import_credit_notes(db: Session, base_path: str) -> None:
    logger.info("Importing credit notes …")
    _truncate(db, CreditNote)

    po_lookup = _build_invoice_po_lookup(db)

    cn_dir = os.path.join(base_path, "Zoho Exports and Proposal", "Credit Notes")
    files = [
        os.path.join(cn_dir, "CN_IBA.xlsx"),
        os.path.join(cn_dir, "CN_NMSA.xlsx"),
    ]

    total = 0
    batch: list[CreditNote] = []
    for fp in files:
        if not os.path.exists(fp):
            logger.warning("CN file not found, skipping: %s", fp)
            continue
        n = _import_cn_file(fp, po_lookup, batch)
        total += n
        if len(batch) >= BATCH_SIZE:
            _flush_batch(db, batch)

    _flush_batch(db, batch)
    logger.info("Credit notes imported: %d rows", total)


# ── Proposals & Pipeline ─────────────────────────────────────────────────────


def import_proposals(db: Session, base_path: str) -> None:
    logger.info("Importing proposals …")
    wb_path = os.path.join(
        base_path, "Zoho Exports and Proposal", "Proposal FY 25-26_Automation.xlsx"
    )
    wb = _open_workbook(wb_path)

    # ── Proposals sheet ──
    try:
        ws = wb["Proposals"]
        prop_rows = list(ws.iter_rows(min_row=2))
    except KeyError:
        logger.warning("'Proposals' sheet not found in proposal workbook")
        prop_rows = []

    # ── Pipeline sheet ──
    try:
        ws2 = wb["Pipeline"]
        pipe_rows = list(ws2.iter_rows(min_row=2))
    except KeyError:
        logger.warning("'Pipeline' sheet not found in proposal workbook")
        pipe_rows = []

    wb.close()

    _truncate(db, Proposal)
    _truncate(db, PipelineEntry)

    # Proposals
    batch: list = []
    count = 0
    for row in prop_rows:
        cust = _str(_cell(row, 9))
        serial = _int(_cell(row, 2))
        if cust is None and serial is None:
            continue

        proposal_status = _str(_cell(row, 14), 50)
        period_str = _str(_cell(row, 7), 100)
        period_date = _parse_date(period_str)
        days_aging: Optional[int] = None
        stale = False
        if period_date is not None:
            days_aging = (date.today() - period_date).days
            stale = (days_aging > 30) and (proposal_status or "").lower() not in ("accepted", "rejected")

        p = Proposal(
            serial_no=serial,
            sub_no=_str(_cell(row, 3), 10),
            year=_str(_cell(row, 4), 20),
            month=_str(_cell(row, 5), 20),
            week=_str(_cell(row, 6), 50),
            period=period_str,
            billing_entity=_str(_cell(row, 8), 50),
            customer_name=_str(cust, 500),
            service_description=_str(_cell(row, 10)),
            service_category=_str(_cell(row, 11), 300),
            department=_str(_cell(row, 12), 100),
            fee_proposed=_float(_cell(row, 13)),
            status=proposal_status,
            follow_up=_str(_cell(row, 15), 200),
            pic_for_so=_str(_cell(row, 16), 200),
            quotation_no=_str(_cell(row, 17), 100),
            so_number=_str(_cell(row, 18), 100),
            remarks=_str(_cell(row, 19)),
            additional_remarks=_str(_cell(row, 20)),
            quotation_status_zoho=_str(_cell(row, 22), 50),
            zoho_remarks=_str(_cell(row, 23)),
            manager_remark=_str(_cell(row, 24)),
            days_since_proposal=days_aging,
            is_stale=stale,
        )
        batch.append(p)
        count += 1
        if len(batch) >= BATCH_SIZE:
            _flush_batch(db, batch)

    _flush_batch(db, batch)
    logger.info("Proposals imported: %d rows", count)

    # Pipeline
    pipe_batch: list = []
    pipe_count = 0
    for row in pipe_rows:
        client = _str(_cell(row, 5))
        discussion = _str(_cell(row, 6))
        if client is None and discussion is None:
            continue

        pe = PipelineEntry(
            year=_str(_cell(row, 1), 20),
            week=_str(_cell(row, 2), 100),
            period=_str(_cell(row, 3), 100),
            billing_entity=_str(_cell(row, 4), 50),
            client_name=_str(client, 500),
            discussion=discussion,
            department=_str(_cell(row, 7), 100),
            follow_up=_str(_cell(row, 8), 200),
            status=_str(_cell(row, 9), 50),
            remarks=_str(_cell(row, 10)),
        )
        pipe_batch.append(pe)
        pipe_count += 1
        if len(pipe_batch) >= BATCH_SIZE:
            _flush_batch(db, pipe_batch)

    _flush_batch(db, pipe_batch)
    logger.info("Pipeline entries imported: %d rows", pipe_count)


# ── Orchestrator ─────────────────────────────────────────────────────────────


def import_all(db: Session, base_path: str) -> dict[str, str]:
    """
    Run every importer in dependency order and commit once at the end.

    Returns a summary dict mapping each data source to its outcome.
    """
    summary: dict[str, str] = {}
    importers = [
        ("master_data", import_master_data),
        ("budget_sheet", import_budget_sheet),
        ("invoices", import_invoices),
        ("sales_orders", import_sales_orders),
        ("credit_notes", import_credit_notes),
        ("proposals", import_proposals),
    ]

    for name, func in importers:
        try:
            logger.info("── Starting %s import ──", name)
            func(db, base_path)
            summary[name] = "ok"
        except FileNotFoundError as exc:
            logger.error("File not found for %s: %s", name, exc)
            summary[name] = f"skipped – {exc}"
        except Exception as exc:
            logger.exception("Error importing %s", name)
            summary[name] = f"error – {exc}"
            db.rollback()
            raise

    db.commit()
    logger.info("All imports committed. Summary: %s", summary)
    return summary
