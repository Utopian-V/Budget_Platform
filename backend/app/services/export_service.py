"""
Excel export service for the Budget Platform.

Generates downloadable .xlsx files for board reports, variance summaries,
and reconciliation data.
"""

import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from sqlalchemy.orm import Session
from sqlalchemy import func

from app.models.models import BudgetLine, BudgetMonthly, ReconciliationRecord

MONTH_LABELS = [
    "Apr-25", "May-25", "Jun-25", "Jul-25", "Aug-25", "Sep-25",
    "Oct-25", "Nov-25", "Dec-25", "Jan-26", "Feb-26", "Mar-26",
]

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def export_department_variance(db: Session, month: str) -> io.BytesIO:
    rows = (
        db.query(
            BudgetLine.department,
            BudgetLine.client_name,
            BudgetLine.service_category,
            BudgetLine.manager,
            BudgetMonthly.expected,
            BudgetMonthly.actual,
            BudgetMonthly.mtd_variance,
            BudgetMonthly.ytd_variance,
            BudgetMonthly.reason,
            BudgetMonthly.remark,
        )
        .join(BudgetMonthly, BudgetLine.id == BudgetMonthly.budget_line_id)
        .filter(BudgetMonthly.month == month)
        .order_by(BudgetLine.department, BudgetLine.client_name)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Dept Variance {month}"

    headers = ["Department", "Client", "Service Category", "Manager",
               "Expected", "Actual", "MTD Variance", "YTD Variance", "Reason", "Remark"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r.department, r.client_name, r.service_category, r.manager,
            r.expected or 0, r.actual or 0, r.mtd_variance or 0, r.ytd_variance or 0,
            r.reason, r.remark,
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_mtd_ytd(db: Session) -> io.BytesIO:
    rows = (
        db.query(
            BudgetLine.department,
            BudgetMonthly.month,
            BudgetMonthly.month_index,
            func.coalesce(func.sum(BudgetMonthly.expected), 0).label("expected"),
            func.coalesce(func.sum(BudgetMonthly.actual), 0).label("actual"),
        )
        .join(BudgetMonthly, BudgetLine.id == BudgetMonthly.budget_line_id)
        .group_by(BudgetLine.department, BudgetMonthly.month, BudgetMonthly.month_index)
        .order_by(BudgetLine.department, BudgetMonthly.month_index)
        .all()
    )

    dept_data: dict[str, dict] = {}
    for r in rows:
        dept = r.department or "Unknown"
        if dept not in dept_data:
            dept_data[dept] = {"dept": dept, "months": {}}
        dept_data[dept]["months"][r.month] = {"expected": r.expected, "actual": r.actual}

    wb = Workbook()
    ws = wb.active
    ws.title = "MTD-YTD Summary"

    header = ["Department"]
    for m in MONTH_LABELS:
        header.extend([f"{m} Exp", f"{m} Act"])
    header.extend(["YTD Expected", "YTD Actual", "YTD Variance"])
    ws.append(header)
    _style_header(ws, 1, len(header))

    for dept, info in dept_data.items():
        row_data = [dept]
        ytd_exp = 0
        ytd_act = 0
        for m in MONTH_LABELS:
            md = info["months"].get(m, {})
            exp = md.get("expected", 0)
            act = md.get("actual", 0)
            ytd_exp += exp
            ytd_act += act
            row_data.extend([exp, act])
        row_data.extend([ytd_exp, ytd_act, ytd_exp - ytd_act])
        ws.append(row_data)

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_client_summary(db: Session) -> io.BytesIO:
    rows = (
        db.query(
            BudgetLine.client_name,
            BudgetLine.department,
            func.coalesce(func.sum(BudgetMonthly.expected), 0).label("total_expected"),
            func.coalesce(func.sum(BudgetMonthly.actual), 0).label("total_actual"),
        )
        .join(BudgetMonthly, BudgetLine.id == BudgetMonthly.budget_line_id)
        .group_by(BudgetLine.client_name, BudgetLine.department)
        .order_by(func.sum(BudgetMonthly.expected).desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Summary"

    headers = ["Client", "Department", "Total Expected", "Total Actual", "Variance"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r.client_name,
            r.department,
            r.total_expected,
            r.total_actual,
            r.total_expected - r.total_actual,
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def export_reconciliation(db: Session, month: str) -> io.BytesIO:
    records = (
        db.query(ReconciliationRecord)
        .filter(ReconciliationRecord.month == month)
        .order_by(ReconciliationRecord.is_matched.desc(), ReconciliationRecord.unique_code)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = f"Reconciliation {month}"

    headers = ["Unique Code", "Budget Amount", "Invoice Amount", "Difference",
               "Matched", "Discrepancy Type", "Detail"]
    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for r in records:
        ws.append([
            r.unique_code,
            r.budget_amount or 0,
            r.invoice_amount or 0,
            r.difference or 0,
            "Yes" if r.is_matched else "No",
            r.discrepancy_type or "",
            r.discrepancy_detail or "",
        ])

    _auto_width(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
